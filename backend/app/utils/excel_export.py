import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime
from typing import List
from io import BytesIO
from app.models.scan import Scan


class ExcelExporter:
    @staticmethod
    def export_scans_to_excel(scans: List[Scan]) -> bytes:
        """Экспортирует сканы в Excel файл"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "История сканирований"

        # Заголовки
        headers = [
            "ID",
            "Данные QR",
            "Тип сканирования",
            "Время сканирования",
            "Напечатано",
            "Время печати",
            "Принтер",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Данные
        for row, scan in enumerate(scans, 2):
            ws.cell(row=row, column=1, value=scan.id)
            ws.cell(row=row, column=2, value=scan.qr_data)
            ws.cell(row=row, column=3, value=scan.scan_type)
            ws.cell(
                row=row, column=4, value=scan.scanned_at.strftime("%Y-%m-%d %H:%M:%S")
            )
            ws.cell(row=row, column=5, value="Да" if scan.printed else "Нет")
            ws.cell(
                row=row,
                column=6,
                value=(
                    scan.printed_at.strftime("%Y-%m-%d %H:%M:%S")
                    if scan.printed_at
                    else ""
                ),
            )
            ws.cell(row=row, column=7, value=scan.printer_id or "")

        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Сохраняем в bytes
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer.getvalue()
